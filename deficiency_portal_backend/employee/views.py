from rest_framework.views import APIView
from django.http import Http404
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated

from school.models import EmployeeProfile, StudentProfile, College
from deficiency.models import Deficiency, FinanceDeficiency
from deficiency.serializers import DeficiencyDetailSerializer, DeficiencyNameListSerializer
from employee.serializers import StudentListSerializer, ReportSerializer, GeneralSummarySerializer, GeneralSummary, PerDeficiencySummarySerializer, PerDeficiencySummary, DashboardDeficiencyNameTableSerializer, BarChartSerializer, BarChartData
from accounts.permissions import HasEmployeePermission
from student.serializers import StudentSummarySerializer
from school.serializers import ProfileSerializer
from django.db.models import Count

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from deficiency_portal_backend.settings import BASE_DIR
import os

# Create your views here.
class DeficiencyDetail(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]

    def get_object(self, def_id):
        try:
            return Deficiency.objects.get(id=def_id)
        except Deficiency.DoesNotExist:
            raise Http404

    def get(self, request, def_id, format=None):
        deficiency_query = self.get_object(def_id)
        
        serializer = DeficiencyDetailSerializer(deficiency_query)
        return Response(serializer.data)
    
    def put(self, request, def_id, format=None):
        deficiency = self.get_object(def_id)
        user = request.user

        deficiency.is_complete = request.data["is_complete"]
        employee = EmployeeProfile.objects.get(employee_id=user.username)

        if not deficiency.is_complete:
            deficiency.processed_by = None
            deficiency.date_fulfilled = None

        else:
            deficiency.processed_by = employee
            deficiency.date_fulfilled = datetime.date.today()
        
        deficiency.save()

        serializer = DeficiencyDetailSerializer(deficiency)
        return Response(serializer.data)
    
    def delete(self, request, def_id, format=None):
        deficiency = self.get_object(def_id)

        try:
            finance_deficiency = FinanceDeficiency.objects.get(deficiency__id=def_id)
            finance_deficiency.delete()
        except FinanceDeficiency.DoesNotExist:
            finance_deficiency = None
        
        deficiency.delete()

        return Response({"success": "The Deficiency was successfully deleted"})

class DeficiencyNameList(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]

    def get(self, request, format=None):
        name = request.GET.get('name')
        
        if name:
            deficiency_name_list = Deficiency.objects.values("name", "category").distinct().filter(name__icontains=name)
        else:
            deficiency_name_list = Deficiency.objects.values("name", "category").distinct()

        serializer = DeficiencyNameListSerializer(deficiency_name_list, many=True)
        return Response(serializer.data)

class StudentList(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, name, format=None):
        student_name = request.GET.get('student-name')
        student_id = request.GET.get('student-id')
        deficiency_name = name

        if not student_name and not student_id:
            print("hello")
            student_list_query = Deficiency.objects.filter(name=deficiency_name).order_by("is_complete")
        
        if student_name and not student_id:
            query_first_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__first_name__icontains=student_name).order_by("is_complete")
            query_last_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__last_name__icontains=student_name).order_by("is_complete")
            query_middle_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__middle_name__icontains=student_name).order_by("is_complete")
            student_list_query = query_first_name | query_last_name | query_middle_name

        if student_id and not student_name:
            student_list_query = Deficiency.objects.filter(name=deficiency_name).filter(student__student_id__icontains=student_id).order_by("is_complete")

        if student_id and student_name:
            print("hello")
            query_first_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__first_name__icontains=student_name).order_by("is_complete")
            query_last_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__last_name__icontains=student_name).order_by("is_complete")
            query_middle_name = Deficiency.objects.filter(name=deficiency_name).filter(student__user__middle_name__icontains=student_name).order_by("is_complete")
            id_query = Deficiency.objects.filter(name=deficiency_name).filter(student__student_id__icontains=student_id).order_by("is_complete")


            student_list_query = (query_first_name | query_last_name | query_middle_name) & id_query

        
        if not student_list_query:
            return Response({"warning": "There are no students with this deficiency yet"})

        serializer = StudentListSerializer(student_list_query, many=True)
        return Response(serializer.data)
    
    def post(self, request, name, *args, **kwargs):
        user = request.user
        student = StudentProfile.objects.get(student_id=request.data["student_id"])
        employee = EmployeeProfile.objects.get(employee_id=user.username)
        

        new_deficiency = Deficiency(student=student, added_by=employee, name=name, category=request.data["category"])
        new_deficiency.save()

        if request.data["category"] == "Finance":
            try:
                amount = request.data["amount"]
                new_finance_deficiency = FinanceDeficiency(deficiency_id=new_deficiency.id, amount=amount)
                
                new_finance_deficiency.save()
            except KeyError:
                return Response({"warning": "no amount given"})

        serializer = DeficiencyDetailSerializer(new_deficiency)
        return Response(serializer.data)

class DeficiencyNameOptions(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, format=None):
        name = request.GET.get('name')
        
        if name:
            deficiency_names = Deficiency.objects.filter(name__icontains=name).order_by('name')
        else:
            deficiency_names = Deficiency.objects.all().order_by('name')

        serializer = DeficiencyNameListSerializer(deficiency_names, many=True)

        return Response(serializer.data)
    
class AllStudents(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, name, format=None):
        student_name = request.GET.get('student-name')
        student_id = request.GET.get('student-id')
        deficiency_name = name

        if not student_name and not student_id:
            student_list_query = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)])
        
        if student_name and not student_id:
            query_first_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__first_name__icontains=student_name)
            query_last_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__last_name__icontains=student_name)
            query_middle_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__middle_name__icontains=student_name)
            student_list_query = query_first_name | query_last_name | query_middle_name

        if student_id and not student_name:
            student_list_query = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(student_id__icontains=student_id)

        if student_id and student_name:
            print("hello")
            query_first_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__first_name__icontains=student_name)
            query_last_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__last_name__icontains=student_name)
            query_middle_name = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(user__middle_name__icontains=student_name)
            id_query = StudentProfile.objects.all().exclude(student_id__in=[x.student.student_id for x in Deficiency.objects.filter(name=deficiency_name)]).filter(student_id__icontains=student_id)

            student_list_query = (query_first_name | query_last_name | query_middle_name) & id_query
        
        if not student_list_query:
            return Response({"warning": "Something went wrong"})

        serializer = StudentSummarySerializer(student_list_query, many=True)
        return Response(serializer.data)
    
class EmployeeProfileView(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]

    def get(self, request,format=None):
        user = self.request.user

        serializer = ProfileSerializer(user)

        return Response(serializer.data)
    
    def put(self, request, format=None):
        user = request.user
        new_number = request.data["mobile_number"]
        new_email = request.data["email"]

        user.mobile_number = new_number
        user.email = new_email

        user.save()

        return Response({"success": "Profile was successfully updated"})
    
class GenerateReportView(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, deficiency_name, format=None):
        file_path = os.path.join(BASE_DIR, "format.xlsx")
        wb = load_workbook(file_path)
        ws = wb.active

        deficiencies = Deficiency.objects.filter(name=deficiency_name).order_by("is_complete")
        serializer = ReportSerializer(deficiencies, many=True)

        if "Balance To Be Settled" in [*serializer.data[0].keys()]:
            ws["O2"] = "Balance To Be Settled"

        # Append Data to excel file
        [ws.append([*x.values()]) for x in serializer.data]

        # Return excel file 
        with NamedTemporaryFile() as tmp:
            wb.save(tmp)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename={deficiency_name} Report.xlsx'
        return response

class GeneralSummaryView(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, format=None):
        summary = GeneralSummary()
        serializer = GeneralSummarySerializer(summary)

        return Response(serializer.data)

class PerDeficiencySummaryView(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, deficiency_name, format=None):
        summary = PerDeficiencySummary(deficiency_name)
        serializer = PerDeficiencySummarySerializer(summary)

        return Response(serializer.data)


class DashboardDeficiencyNameTable(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, format=None):
        name = request.GET.get('name')

        try:
            if name:
                deficiency_name_list = Deficiency.objects.values("name", "category").distinct().filter(name__icontains=name)
            else:
                deficiency_name_list = Deficiency.objects.values("name", "category").distinct()

        except Exception as e:
            return Response({"warning" : e})

        serializer = DashboardDeficiencyNameTableSerializer(deficiency_name_list, many=True)
        return Response(serializer.data)
    
class BarChartSummary(APIView):
    permission_classes = [IsAuthenticated & HasEmployeePermission]
    def get(self, request, format=None):
        name = request.GET.get('name')

        try:
            if name:
                top_5_colleges = College.objects.annotate(def_count=Count("department__studentprofile__student_with_deficiency")).filter(department__studentprofile__student_with_deficiency__name=name).order_by("-def_count")[:5]
            else:
                top_5_colleges = College.objects.annotate(def_count=Count("department__studentprofile__student_with_deficiency")).order_by("-def_count")[:5]
            
        except Exception as e:
            return Response({"warning" : e})
        
        serializer = BarChartSerializer([BarChartData(x.college_abbreviation, name) for x in top_5_colleges], many=True)
        
        return Response(serializer.data)